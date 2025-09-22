import openpyxl

def rowcount(file,sheetname): #take parameter from ddt test case file and send to open-pyexcel
    workbook=openpyxl.load_workbook(file) #to load the workbook we use this statement
    sheet=workbook[sheetname]
    return(sheet.max_row)

def columncount(file,sheetname): #file is the excel file name, sheet name is the sheet in which we are working
    workbook=openpyxl.load_workbook(file) #loads our file
    sheet=workbook[sheetname] #loads the working sheet of our file
    return (sheet.max_column)

def readdata(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value

def writedata(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(file)